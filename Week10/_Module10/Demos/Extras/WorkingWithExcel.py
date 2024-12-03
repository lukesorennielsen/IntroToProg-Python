'''*************************************************************************--
-- Desc: This script reads data from an Excel file
-- Change Log: When,Who,What
-- 2020-05-15,RRoot,Created File
--**************************************************************************'''
# Verify the following setup!
# Python -m pip list
# Python -m pip install openpyxl


from openpyxl import load_workbook

# Indicate where the file is
strFileName = 'MyData.xlsx'

# Use the constructor to open the Excel file
# and store it's data in the records variable
objWorkbook = load_workbook(filename=strFileName)
objSheet = objWorkbook.active

# Print a header to make the results more attractive
print('CategoryName', 'ProductName', 'OrderQtyTotal', 'OrderPriceTotal', sep=',')

# Get rows of data from the spreadsheet
for row in objSheet.iter_rows(min_row=2, values_only= True):
    print(row[0],row[1],row[2],row[3],sep=',')

print(' Produce Only ----------------------------------')
print('ProductName', 'OrderQtyTotal', 'OrderPriceTotal', sep=',')
for row in objSheet.iter_rows(min_row=2, values_only= True):
    if row[0] == 'Seafood':  # Only show seafood data for short results
        print(row[1],row[2],row[3],sep=',')



# Important: Make sure to watch this video for info on OpenPyXl https://youtu.be/_uQrJ0TkZlc?t=14158
# The video has some great example of how to work within the spreadsheet
# OpenPyXl is another way to work with Excel files
